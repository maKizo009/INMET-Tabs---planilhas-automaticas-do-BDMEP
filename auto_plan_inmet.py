from cProfile import label
from tkinter import filedialog
import pandas as pd
import csv
import os
from openpyxl import load_workbook
from shutil import copyfile
import tkinter as tk

def encerrar_aplicativo():
    root.destroy()

root = tk.Tk()
root.withdraw()

caminho_para_downloads = os.path.expanduser("~/Downloads")
arquivos_csv = filedialog.askopenfilenames(initialdir=caminho_para_downloads,
                                           title="Selecione os arquivos", 
                                           filetypes=(("csv files", "*.csv"), ("all files", "*.*")))

janela = tk.Toplevel(root)
janela.title("Processando arquivo")
label = tk.Label(janela, text="")
label.pack()
botao_encerrar = tk.Button(janela, text="Encerrar", command=encerrar_aplicativo)
botao_encerrar.pack()

for caminho_completo_para_arquivo in arquivos_csv:
    with open(caminho_completo_para_arquivo, "r") as f_entrada:
        leitor = csv.reader(f_entrada, delimiter=";")

        primeira_linha = f_entrada.readline().rstrip()

        label.config(text=f"Dados da estação de: {primeira_linha[6:]}. Não feche essa janela ou a planilha poderá ser corrompida")
        janela.update()

        leitor = pd.DataFrame(leitor)

        leitor = leitor.iloc[11:]

        leitor = leitor.rename(columns={0: "Data", 1: "Max", 2: "Min", 3: "Precipitação"})
        leitor = leitor.reindex(columns=["Data", "Min", "Max"])

        leitor["Max"] = pd.to_numeric(leitor["Max"], errors="coerce")
        leitor["Min"] = pd.to_numeric(leitor["Min"], errors="coerce")
        
        linha = leitor["Max"].idxmin()
        colunas = ["Data", "Min", "Max"]
        resultado = leitor.loc[linha, colunas]

        linha2 = leitor["Min"].idxmin()
        colunas2 = ["Data", "Min", "Max"]
        resultado2 = leitor.loc[linha2, colunas2]

        linha3 = leitor["Max"].idxmax()
        colunas3 = ["Data", "Min", "Max"]
        resultado3 = leitor.loc[linha3, colunas3]
        
        linha4 = leitor["Min"].idxmax()
        colunas4 = ["Data", "Min", "Max"]
        resultado4 = leitor.loc[linha4, colunas4]

        leitor["Data"] = pd.to_datetime(leitor["Data"])

        leitor["Dia"] = leitor["Data"].dt.day
        leitor["Mes"] = leitor["Data"].dt.month
        leitor["Ano"] = leitor["Data"].dt.year

    caminhoOriginal = os.path.expanduser("~\\Downloads\\Modelo para normais climatológicas 1991-2020.xlsx")
    novoCaminho = os.path.expanduser(f'~\\Documents\\{primeira_linha[6:]}_BDMEP.xlsx')

    copyfile(caminhoOriginal, novoCaminho)

    book = load_workbook(caminhoOriginal)
    
    for ano in leitor["Ano"].dropna().unique():
        dadosAno = leitor[leitor["Ano"]== ano]
        organizado = pd.DataFrame()

        for mes in range(1, 13):
            dadosMes = dadosAno[dadosAno["Mes"] == mes]
            minimas = dadosMes[["Dia", "Min"]].set_index("Dia")
            maximas = dadosMes[["Dia", "Max"]].set_index("Dia")
            minimas.columns = [f"Minima{mes}"]
            maximas.columns = [f"Máxima{mes}"]
            organizado.index = organizado.index.astype(str)
            inicio = pd.to_datetime(f'{int(ano)}-{int(mes):02d}-01')
            fim = inicio + pd.offsets.MonthEnd(0)
            dias = pd.date_range(inicio, fim)
            dias = dias.day
            minimas = minimas.reindex(dias)
            maximas = maximas.reindex(dias)
            organizado = organizado.reset_index(drop=True)
            minimas = minimas.reset_index(drop=True)
            maximas = maximas.reset_index(drop=True)
            organizado = pd.concat([organizado, minimas, maximas], axis=1)

        sheet_name = str(int(ano))
        if sheet_name in book.sheetnames:
            ws = book[sheet_name]
        else:
            ws = book.create_sheet(sheet_name)

        for col, col_data in enumerate(organizado.values.T, start=2):
            for row, value in enumerate(col_data, start=3):
                if pd.isna(value):
                    value = "-"
                ws.cell(row=row, column=col, value=value)

    book.save(os.path.expanduser(novoCaminho))
    label.config(text=f"Tabela da estação de {primeira_linha[6:]} salva na pasta Documentos. Você agora pode encerrar essa aplicação")
janela.update()
root.mainloop()

